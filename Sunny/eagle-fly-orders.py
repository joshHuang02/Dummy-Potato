# 鹰飞打单 - 这个程序因为eccang限制只处理FedEx单，UPS单需要手动输入, 只测试过amazon, ebay, shopify订单。其他平台需要手动输入
# assumes original csv is sorted already, check that 2 piece sets are ordered correctly nexst to each other
# use pyinsatller to translate python program to .exe, use command 'pyinstaller --onefile --distpath ./ --clean -y -w eagle-fly-orders.py', this will replace the old .exe
import csv
import sys
import tkinter.filedialog as fd
from datetime import datetime
from tkinter import Tk, filedialog, messagebox

import openpyxl
from openpyxl import Workbook


# stops the program with an error message in case anything goes wrong
def terminate(errmsg):
    messagebox.showinfo("error", errmsg)
    sys.exit()

# removes extra characters "  =  from csv value, also removes everything after 'ext' to format phone numbers
def clean(value, separator='ext'):
    return value.replace('"', "").replace('=', "").split(separator, 1)[0]

# processes the input csv files from eccang, selects the orders that match the 鹰飞 ordes and adds them to the 出库 data template
def process_csv(待处理列表, 出库数据):
    # check if list is empty, skip if empty
    if (待处理列表 == ""): return
    
    with open(待处理列表, mode='rt', encoding="utf8") as csv_file:
        csv_reader = list(csv.DictReader(csv_file))
        for row in csv_reader:
            出库数据.line_count += 1
            #last 5 characters in reference number (could be characters if shopify)
            reference_no = clean(row["refrence_no"])[-5:]
            # if a order number matches
            if (reference_no in order_list):
                output_row = ()
                # if order is multiple of many parts, following parts only need sku and qty
                if (reference_no in duplicate_orders_list):
                    output_row = (row["pcr_product_sku"], row["qty"])
                else:
                    # add the unique order to duplicates list to check for future duplicates
                    duplicate_orders_list.append(reference_no)
                    # sku, qty, name, , phone, street, , , city, state, zip, , country code, , , refrence no, , , , , , , , , tracking no
                    output_row = (row["pcr_product_sku"], row["qty"], row["consignee_name"], "", row["consignee_phone"], row["consignee_street"], "", "", row["consignee_city"],
                                    row["consignee_province"], row["consignee_zip"], "", row["consignee_country_code"], "", "", row["refrence_no"], "", "", "", "", "", "", "", "", row["tracking_number"])

                # cleans the data for formatting by converting to list since tuple is immutable
                output_row = list(output_row)
                for i in range(len(output_row)):
                    output_row[i] = clean(output_row[i])
                # append the new lines of data to output xlsx file and updates quantity count for number of labels found
                出库数据.output_xlsx.append(output_row)
                出库数据.label_count += int(clean(row["qty"]))

# centuralized object for storing the output information
class Output:
    def __init__(self):
        self.line_count = 0
        self.label_count = 0
        self.output_xlsx = book.active


# initialte global variables and objects
order_list = []
duplicate_orders_list = []
book = Workbook()
出库数据 = Output()

# main code
def main():
    # select necessary files, terminate program if first one is not selected, otherwise continue
    Tk().withdraw()
    出库单列表 = fd.askopenfilename(title='出库单', filetypes=[('Excel files', '.xlsx .xls')])
    if (出库单列表 == ""): terminate("未选择出库单列表")
    待处理单列表 = list(fd.askopenfilenames(title='选择所有待处理单', filetypes=[('CSV', '.csv')]))

    # generate list of order numbers, should be xlsx sheet of last 5 characters of order number
    出库单列表 = openpyxl.load_workbook(出库单列表).active
    for row in 出库单列表.iter_rows():
        for cell in row:
            order_list.append(str(cell.value))
    if (len(order_list) == 0): terminate("出库单列表没有内容，请确认保存过列表")

    # process all imported csv files
    for 待处理单 in 待处理单列表:
        process_csv(待处理单, 出库数据)

    # save output file to selected folder with auto filename including time
    now = datetime.now()
    current_time = now.strftime("%m-%d-%Y_%H-%M-%S")
    save_location = filedialog.asksaveasfilename(title='出货数据文件夹', defaultextension='.xlsx', initialfile=f'出库数据 {current_time}')
    if (save_location == ""): terminate("未选择保存地点")
    book.save(save_location)

    messagebox.showinfo("完成！", f'处理{len(待处理单列表)}个店面的{出库数据.line_count}张单，出库{出库数据.label_count}张单')

if __name__ == "__main__":
    main()